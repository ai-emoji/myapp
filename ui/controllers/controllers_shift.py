# controllers_shift.py
# Controller cho widgets_shift.py

import traceback
from PySide6.QtWidgets import QMessageBox, QTableWidgetItem, QProgressDialog, QInputDialog, QCheckBox
from PySide6.QtCore import Qt, QThread, Signal
from datetime import datetime


def log_to_debug(message):
    try:
        with open("log/debug.log", "a", encoding="utf-8") as f:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{now}] {message}\n")
    except Exception as e:
        print(f"[LogError] {e}")


class UploadThread(QThread):
    """Thread để tải nhân viên lên máy trong background"""
    progress = Signal(int, str)  # progress value, status message
    finished_signal = Signal(bool, str, int)  # success, message, count
    
    def __init__(self, device_id, employee_ids):
        super().__init__()
        self.device_id = device_id
        self.employee_ids = employee_ids
    
    def run(self):
        try:
            from services.shift_upload_services import ShiftUploadService
            
            self.progress.emit(0, "Đang chuẩn bị...")
            
            service = ShiftUploadService()
            
            def progress_callback(progress_value, message):
                self.progress.emit(progress_value, message)
            
            success, message, count = service.upload_employees_to_device(
                self.device_id, 
                self.employee_ids,
                progress_callback
            )
            
            self.progress.emit(100, "Hoàn tất!")
            self.finished_signal.emit(success, message, count)
            
        except Exception as e:
            log_to_debug(f"UploadThread error: {e}\n{traceback.format_exc()}")
            self.finished_signal.emit(False, f"Lỗi: {str(e)}", 0)


class DeleteThread(QThread):
    """Thread để xóa nhân viên khỏi máy trong background"""
    progress = Signal(int, str)
    finished_signal = Signal(bool, str, int)
    
    def __init__(self, device_id, user_ids, delete_type="employee"):
        super().__init__()
        self.device_id = device_id
        self.user_ids = user_ids
        self.delete_type = delete_type  # "employee" hoặc "fingerprint"
    
    def run(self):
        try:
            from services.shift_upload_services import ShiftUploadService
            
            self.progress.emit(0, "Đang chuẩn bị...")
            
            service = ShiftUploadService()
            
            def progress_callback(progress_value, message):
                self.progress.emit(progress_value, message)
            
            if self.delete_type == "fingerprint":
                success, message, count = service.delete_fingerprints_from_device(
                    self.device_id,
                    self.user_ids,
                    progress_callback
                )
            else:
                success, message, count = service.delete_employees_from_device(
                    self.device_id,
                    self.user_ids,
                    progress_callback
                )
            
            self.progress.emit(100, "Hoàn tất!")
            self.finished_signal.emit(success, message, count)
            
        except Exception as e:
            log_to_debug(f"DeleteThread error: {e}\n{traceback.format_exc()}")
            self.finished_signal.emit(False, f"Lỗi: {str(e)}", 0)


class ControllerWidgetsShift:
    """Controller xử lý logic cho widgets_shift.py"""

    def __init__(self, widget):
        self.widget = widget
        self.current_device_id = None
        
        # Initialize services
        from services.shift_upload_services import ShiftUploadService
        from services.employee_services import EmployeeService
        from services.device_services import DeviceService
        
        self.shift_service = ShiftUploadService()
        self.employee_service = EmployeeService()
        self.device_service = DeviceService()
        
        # Connect signals
        self._connect_signals()
        
        # Load initial data
        self._load_employees()

    def _connect_signals(self):
        """Kết nối các signals"""
        try:
            # Employee section
            self.widget.txt_search_employee.textChanged.connect(self._on_search_employee)
            self.widget.btn_refresh_employees.clicked.connect(self._load_employees)
            self.widget.btn_download_from_device.clicked.connect(self._move_to_upload_list)
            self.widget.btn_upload_file.clicked.connect(self._upload_employee_list)
            
            # Uploaded section
            self.widget.btn_remove.clicked.connect(self._remove_from_upload_list)
            self.widget.btn_select_upload.clicked.connect(self._select_device_and_upload)
            self.widget.btn_select_delete.clicked.connect(self._select_device_and_delete)
            self.widget.btn_select_delete_fp.clicked.connect(self._select_device_and_delete_fingerprint)
            
            log_to_debug("ControllerWidgetsShift: Signals connected")
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _connect_signals() error: {e}")

    def _load_employees(self):
        """Tải danh sách nhân viên"""
        try:
            log_to_debug("ControllerWidgetsShift: _load_employees() called")
            
            employees = self.employee_service.get_all_employees()
            
            self.widget.table_employees.setRowCount(0)
            self.widget.table_employees.setSortingEnabled(False)
            
            for emp in employees:
                row = self.widget.table_employees.rowCount()
                self.widget.table_employees.insertRow(row)
                
                # Column 0: Checkbox
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 10px;")
                self.widget.table_employees.setCellWidget(row, 0, checkbox)
                
                # Store employee_id in checkbox
                checkbox.setProperty("employee_id", emp["id"])
                
                # Column 1: Mã nhân viên
                item = QTableWidgetItem(emp.get("employee_code", ""))
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 1, item)
                
                # Column 2: Tên nhân viên
                item = QTableWidgetItem(emp.get("name", ""))
                self.widget.table_employees.setItem(row, 2, item)
                
                # Column 3: Mã chấm công
                item = QTableWidgetItem(emp.get("attendance_code", ""))
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 3, item)
                
                # Column 4: Tên chấm công
                item = QTableWidgetItem(emp.get("attendance_name", ""))
                self.widget.table_employees.setItem(row, 4, item)
                
                # Column 5: Mã số thẻ (placeholder)
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 5, item)
                
                # Column 6: Mật mã (placeholder)
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 6, item)
                
                # Column 7: Loại (0=User)
                item = QTableWidgetItem("0")
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 7, item)
                
                # Column 8: Cho phép
                item = QTableWidgetItem("✅")
                item.setTextAlignment(Qt.AlignCenter)
                self.widget.table_employees.setItem(row, 8, item)
            
            self.widget.table_employees.setSortingEnabled(True)
            self.widget.lbl_total_employees.setText(f"Tổng số: {len(employees)}")
            
            log_to_debug(f"ControllerWidgetsShift: Loaded {len(employees)} employees")
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _load_employees() error: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self.widget, "Lỗi", f"Không thể tải danh sách nhân viên: {str(e)}")

    def _on_search_employee(self, text):
        """Tìm kiếm nhân viên"""
        try:
            search_text = text.strip().lower()
            
            visible_count = 0
            for row in range(self.widget.table_employees.rowCount()):
                # Get employee code and name
                code_item = self.widget.table_employees.item(row, 1)
                name_item = self.widget.table_employees.item(row, 2)
                att_code_item = self.widget.table_employees.item(row, 3)
                att_name_item = self.widget.table_employees.item(row, 4)
                
                code = code_item.text().lower() if code_item else ""
                name = name_item.text().lower() if name_item else ""
                att_code = att_code_item.text().lower() if att_code_item else ""
                att_name = att_name_item.text().lower() if att_name_item else ""
                
                # Show row if match
                if (search_text == "" or search_text in code or search_text in name or 
                    search_text in att_code or search_text in att_name):
                    self.widget.table_employees.setRowHidden(row, False)
                    visible_count += 1
                else:
                    self.widget.table_employees.setRowHidden(row, True)
            
            log_to_debug(f"Search '{search_text}': {visible_count} rows visible")
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_search_employee() error: {e}")

    def _move_to_upload_list(self):
        """Chuyển nhân viên đã chọn xuống danh sách tải lên"""
        try:
            # Get selected employees (checked checkboxes)
            selected_employees = []
            
            for row in range(self.widget.table_employees.rowCount()):
                checkbox = self.widget.table_employees.cellWidget(row, 0)
                if checkbox and checkbox.isChecked():
                    employee_id = checkbox.property("employee_id")
                    
                    # Get employee data
                    emp_data = {
                        "employee_id": employee_id,
                        "employee_code": self.widget.table_employees.item(row, 1).text(),
                        "name": self.widget.table_employees.item(row, 2).text(),
                        "attendance_code": self.widget.table_employees.item(row, 3).text(),
                        "attendance_name": self.widget.table_employees.item(row, 4).text(),
                    }
                    selected_employees.append(emp_data)
            
            if not selected_employees:
                QMessageBox.warning(
                    self.widget, 
                    "Cảnh báo", 
                    "Vui lòng chọn ít nhất một nhân viên"
                )
                return
            
            # Add to uploaded table
            for emp in selected_employees:
                # Check if already exists
                existing = False
                for row in range(self.widget.table_uploaded.rowCount()):
                    code_item = self.widget.table_uploaded.item(row, 0)
                    if code_item and code_item.text() == emp["employee_code"]:
                        existing = True
                        break
                
                if not existing:
                    row = self.widget.table_uploaded.rowCount()
                    self.widget.table_uploaded.insertRow(row)
                    
                    # Store employee_id in first column
                    item = QTableWidgetItem(emp["employee_code"])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setData(Qt.UserRole, emp["employee_id"])  # Store employee_id
                    self.widget.table_uploaded.setItem(row, 0, item)
                    
                    item = QTableWidgetItem(emp["attendance_code"])
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(row, 1, item)
                    
                    item = QTableWidgetItem(emp["attendance_name"])
                    self.widget.table_uploaded.setItem(row, 2, item)
                    
                    item = QTableWidgetItem("")  # Mã số thẻ
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(row, 3, item)
                    
                    item = QTableWidgetItem("")  # Mật mã
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(row, 4, item)
                    
                    item = QTableWidgetItem("0")  # Loại
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(row, 5, item)
                    
                    item = QTableWidgetItem("✅")  # Cho phép
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(row, 6, item)
            
            # Update count
            self.widget.lbl_total_uploaded.setText(
                f"Tổng số: {self.widget.table_uploaded.rowCount()}"
            )
            
            # Uncheck all checkboxes
            for row in range(self.widget.table_employees.rowCount()):
                checkbox = self.widget.table_employees.cellWidget(row, 0)
                if checkbox:
                    checkbox.setChecked(False)
            
            QMessageBox.information(
                self.widget,
                "Thành công",
                f"Đã thêm {len(selected_employees)} nhân viên vào danh sách tải lên"
            )
            
        except Exception as e:
            log_to_debug(
                f"ControllerWidgetsShift: _move_to_upload_list() error: {e}\n{traceback.format_exc()}"
            )
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _upload_employee_list(self):
        """Upload danh sách nhân viên từ file CSV/Excel"""
        try:
            from PySide6.QtWidgets import QFileDialog
            import csv
            
            # Mở dialog chọn file
            file_path, _ = QFileDialog.getOpenFileName(
                self.widget,
                "Chọn file danh sách nhân viên",
                "",
                "CSV Files (*.csv);;Excel Files (*.xlsx *.xls);;All Files (*.*)"
            )
            
            if not file_path:
                return
            
            log_to_debug(f"ControllerWidgetsShift: Uploading file: {file_path}")
            
            # Đọc file CSV
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                # Xử lý Excel file
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    
                    # Lấy header từ dòng đầu
                    headers = [cell.value for cell in ws[1]]
                    
                    # Đọc dữ liệu
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                        rows.append(row_dict)
                except ImportError:
                    QMessageBox.warning(
                        self.widget,
                        "Cảnh báo",
                        "Chưa cài đặt thư viện openpyxl. Vui lòng chạy: pip install openpyxl"
                    )
                    return
            else:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Định dạng file không được hỗ trợ"
                )
                return
            
            if not rows:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "File rỗng hoặc không có dữ liệu"
                )
                return
            
            # Lấy danh sách tất cả nhân viên để tìm kiếm
            all_employees = self.employee_service.get_all_employees()
            employee_dict_by_code = {emp["employee_code"]: emp for emp in all_employees}
            
            added_count = 0
            not_found = []
            
            # Xử lý từng dòng trong file
            for row in rows:
                # Lấy mã nhân viên từ file (hỗ trợ nhiều tên cột)
                employee_code = (row.get("Mã nhân viên") or 
                               row.get("employee_code") or 
                               row.get("Ma nhan vien") or
                               row.get("Employee Code") or "").strip()
                
                if not employee_code:
                    continue
                
                # Tìm nhân viên trong DB
                emp = employee_dict_by_code.get(employee_code)
                
                if not emp:
                    not_found.append(employee_code)
                    continue
                
                # Kiểm tra đã tồn tại chưa
                existing = False
                for r in range(self.widget.table_uploaded.rowCount()):
                    code_item = self.widget.table_uploaded.item(r, 0)
                    if code_item and code_item.text() == employee_code:
                        existing = True
                        break
                
                if not existing:
                    # Thêm vào bảng uploaded
                    r = self.widget.table_uploaded.rowCount()
                    self.widget.table_uploaded.insertRow(r)
                    
                    item = QTableWidgetItem(emp["employee_code"])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setData(Qt.UserRole, emp["id"])
                    self.widget.table_uploaded.setItem(r, 0, item)
                    
                    item = QTableWidgetItem(emp.get("attendance_code", ""))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(r, 1, item)
                    
                    item = QTableWidgetItem(emp.get("attendance_name") or emp.get("name", ""))
                    self.widget.table_uploaded.setItem(r, 2, item)
                    
                    item = QTableWidgetItem("")  # Mã số thẻ
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(r, 3, item)
                    
                    item = QTableWidgetItem("")  # Mật mã
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(r, 4, item)
                    
                    item = QTableWidgetItem("0")  # Loại
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(r, 5, item)
                    
                    item = QTableWidgetItem("✅")  # Cho phép
                    item.setTextAlignment(Qt.AlignCenter)
                    self.widget.table_uploaded.setItem(r, 6, item)
                    
                    added_count += 1
            
            # Update count
            self.widget.lbl_total_uploaded.setText(
                f"Tổng số: {self.widget.table_uploaded.rowCount()}"
            )
            
            # Thông báo kết quả
            message = f"Đã thêm {added_count} nhân viên vào danh sách tải lên"
            if not_found:
                message += f"\n\nKhông tìm thấy {len(not_found)} mã nhân viên: {', '.join(not_found[:5])}"
                if len(not_found) > 5:
                    message += f"... và {len(not_found) - 5} mã khác"
            
            QMessageBox.information(self.widget, "Kết quả", message)
            
            log_to_debug(f"ControllerWidgetsShift: Uploaded {added_count} employees from file")
            
        except Exception as e:
            log_to_debug(
                f"ControllerWidgetsShift: _upload_employee_list() error: {e}\n{traceback.format_exc()}"
            )
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi khi đọc file: {str(e)}")

    def _remove_from_upload_list(self):
        """Xóa nhân viên khỏi danh sách tải lên"""
        try:
            selected_rows = self.widget.table_uploaded.selectedIndexes()
            
            if not selected_rows:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Vui lòng chọn ít nhất một dòng để xóa"
                )
                return
            
            # Get unique rows
            rows = sorted(set(index.row() for index in selected_rows), reverse=True)
            
            reply = QMessageBox.question(
                self.widget,
                "Xác nhận",
                f"Bạn có chắc muốn xóa {len(rows)} nhân viên khỏi danh sách tải lên?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                for row in rows:
                    self.widget.table_uploaded.removeRow(row)
                
                self.widget.lbl_total_uploaded.setText(
                    f"Tổng số: {self.widget.table_uploaded.rowCount()}"
                )
                
                QMessageBox.information(
                    self.widget,
                    "Thành công",
                    f"Đã xóa {len(rows)} nhân viên khỏi danh sách"
                )
            
        except Exception as e:
            log_to_debug(
                f"ControllerWidgetsShift: _remove_from_upload_list() error: {e}\n{traceback.format_exc()}"
            )
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _select_device_and_upload(self):
        """Chọn máy và tải nhân viên lên"""
        try:
            # Check if there are employees to upload
            if self.widget.table_uploaded.rowCount() == 0:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Danh sách tải lên trống. Vui lòng chọn nhân viên."
                )
                return
            
            # Get list of devices
            devices = self.device_service.get_all_devices()
            
            if not devices:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Không có thiết bị nào. Vui lòng thêm thiết bị trước."
                )
                return
            
            # Create device selection dialog
            device_names = [f"{d['device_number']} - {d['device_name']}" for d in devices]
            
            device_name, ok = QInputDialog.getItem(
                self.widget,
                "Chọn thiết bị",
                "Chọn máy chấm công để tải nhân viên lên:",
                device_names,
                0,
                False
            )
            
            if ok and device_name:
                # Find device by name
                selected_device = None
                for d in devices:
                    if f"{d['device_number']} - {d['device_name']}" == device_name:
                        selected_device = d
                        break
                
                if selected_device:
                    # Get employee IDs from upload table
                    employee_ids = []
                    for row in range(self.widget.table_uploaded.rowCount()):
                        item = self.widget.table_uploaded.item(row, 0)
                        if item:
                            emp_id = item.data(Qt.UserRole)
                            if emp_id:
                                employee_ids.append(emp_id)
                    
                    # Start upload thread
                    self._start_upload(selected_device["id"], employee_ids)
            
        except Exception as e:
            log_to_debug(
                f"ControllerWidgetsShift: _select_device_and_upload() error: {e}\n{traceback.format_exc()}"
            )
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _start_upload(self, device_id, employee_ids):
        """Bắt đầu tải nhân viên lên máy"""
        try:
            # Create progress dialog
            self.progress_dialog = QProgressDialog(
                "Đang tải nhân viên lên máy chấm công...",
                "Hủy",
                0,
                100,
                self.widget
            )
            self.progress_dialog.setWindowTitle("Tải lên máy chấm công")
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setMinimumDuration(0)
            self.progress_dialog.setAutoClose(False)
            self.progress_dialog.setAutoReset(False)
            
            # Gradient style
            self.progress_dialog.setStyleSheet("""
                QProgressDialog {
                    min-width: 400px;
                }
                QProgressBar {
                    border: 2px solid #2C3E50;
                    border-radius: 5px;
                    text-align: center;
                    background: white;
                }
                QProgressBar::chunk {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #3498db, stop:0.5 #2ecc71, stop:1 #27ae60);
                    border-radius: 3px;
                }
            """)
            
            # Create and start thread
            self.upload_thread = UploadThread(device_id, employee_ids)
            self.upload_thread.progress.connect(self._on_upload_progress)
            self.upload_thread.finished_signal.connect(self._on_upload_finished)
            self.upload_thread.start()
            
            self.progress_dialog.canceled.connect(self._on_upload_canceled)
            self.progress_dialog.show()
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _start_upload() error: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _on_upload_progress(self, value, message):
        """Cập nhật tiến trình upload"""
        try:
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                self.progress_dialog.setValue(value)
                self.progress_dialog.setLabelText(message)
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_upload_progress() error: {e}")

    def _on_upload_finished(self, success, message, count):
        """Xử lý khi upload hoàn tất"""
        try:
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                self.progress_dialog.close()
            
            if success:
                QMessageBox.information(self.widget, "Thành công", message)
                # Clear upload list after successful upload
                self.widget.table_uploaded.setRowCount(0)
                self.widget.lbl_total_uploaded.setText("Tổng số: 0")
            else:
                QMessageBox.warning(self.widget, "Thất bại", message)
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_upload_finished() error: {e}")

    def _on_upload_canceled(self):
        """Xử lý khi hủy upload"""
        try:
            if hasattr(self, 'upload_thread') and self.upload_thread:
                self.upload_thread.terminate()
                self.upload_thread.wait()
            
            QMessageBox.information(self.widget, "Đã hủy", "Đã hủy quá trình tải lên")
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_upload_canceled() error: {e}")

    def _select_device_and_delete(self):
        """Chọn máy và xóa nhân viên"""
        try:
            selected_rows = self.widget.table_uploaded.selectedIndexes()
            
            if not selected_rows:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Vui lòng chọn nhân viên để xóa"
                )
                return
            
            # Get devices
            devices = self.device_service.get_all_devices()
            
            if not devices:
                QMessageBox.warning(self.widget, "Cảnh báo", "Không có thiết bị nào")
                return
            
            device_names = [f"{d['device_number']} - {d['device_name']}" for d in devices]
            device_name, ok = QInputDialog.getItem(
                self.widget,
                "Chọn thiết bị",
                "Chọn máy chấm công để xóa nhân viên:",
                device_names,
                0,
                False
            )
            
            if ok and device_name:
                selected_device = None
                for d in devices:
                    if f"{d['device_number']} - {d['device_name']}" == device_name:
                        selected_device = d
                        break
                
                if selected_device:
                    # Get user_ids
                    rows = sorted(set(index.row() for index in selected_rows))
                    user_ids = []
                    
                    for row in rows:
                        item = self.widget.table_uploaded.item(row, 1)  # attendance_code
                        if item:
                            user_ids.append(item.text())
                    
                    self._start_delete(selected_device["id"], user_ids, "employee")
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _select_device_and_delete() error: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _select_device_and_delete_fingerprint(self):
        """Chọn máy và xóa vân tay"""
        try:
            selected_rows = self.widget.table_uploaded.selectedIndexes()
            
            if not selected_rows:
                QMessageBox.warning(
                    self.widget,
                    "Cảnh báo",
                    "Vui lòng chọn nhân viên để xóa vân tay"
                )
                return
            
            # Get devices
            devices = self.device_service.get_all_devices()
            
            if not devices:
                QMessageBox.warning(self.widget, "Cảnh báo", "Không có thiết bị nào")
                return
            
            device_names = [f"{d['device_number']} - {d['device_name']}" for d in devices]
            device_name, ok = QInputDialog.getItem(
                self.widget,
                "Chọn thiết bị",
                "Chọn máy chấm công để xóa vân tay:",
                device_names,
                0,
                False
            )
            
            if ok and device_name:
                selected_device = None
                for d in devices:
                    if f"{d['device_number']} - {d['device_name']}" == device_name:
                        selected_device = d
                        break
                
                if selected_device:
                    # Get user_ids
                    rows = sorted(set(index.row() for index in selected_rows))
                    user_ids = []
                    
                    for row in rows:
                        item = self.widget.table_uploaded.item(row, 1)  # attendance_code
                        if item:
                            user_ids.append(item.text())
                    
                    self._start_delete(selected_device["id"], user_ids, "fingerprint")
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _select_device_and_delete_fingerprint() error: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _start_delete(self, device_id, user_ids, delete_type):
        """Bắt đầu xóa nhân viên/vân tay"""
        try:
            # Create progress dialog
            title = "Xóa vân tay" if delete_type == "fingerprint" else "Xóa nhân viên"
            self.progress_dialog = QProgressDialog(
                f"Đang {title.lower()} khỏi máy chấm công...",
                "Hủy",
                0,
                100,
                self.widget
            )
            self.progress_dialog.setWindowTitle(title)
            self.progress_dialog.setWindowModality(Qt.WindowModal)
            self.progress_dialog.setMinimumDuration(0)
            self.progress_dialog.setAutoClose(False)
            self.progress_dialog.setAutoReset(False)
            
            # Create and start thread
            self.delete_thread = DeleteThread(device_id, user_ids, delete_type)
            self.delete_thread.progress.connect(self._on_delete_progress)
            self.delete_thread.finished_signal.connect(self._on_delete_finished)
            self.delete_thread.start()
            
            self.progress_dialog.show()
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _start_delete() error: {e}\n{traceback.format_exc()}")
            QMessageBox.critical(self.widget, "Lỗi", f"Lỗi: {str(e)}")

    def _on_delete_progress(self, value, message):
        """Cập nhật tiến trình xóa"""
        try:
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                self.progress_dialog.setValue(value)
                self.progress_dialog.setLabelText(message)
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_delete_progress() error: {e}")

    def _on_delete_finished(self, success, message, count):
        """Xử lý khi xóa hoàn tất"""
        try:
            if hasattr(self, 'progress_dialog') and self.progress_dialog:
                self.progress_dialog.close()
            
            if success:
                QMessageBox.information(self.widget, "Thành công", message)
            else:
                QMessageBox.warning(self.widget, "Thất bại", message)
            
        except Exception as e:
            log_to_debug(f"ControllerWidgetsShift: _on_delete_finished() error: {e}")
