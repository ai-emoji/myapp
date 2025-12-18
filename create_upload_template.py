# Script tạo file Excel mẫu upload nhân viên
# Chạy script này để tạo file upload_nhanvien_mau.xlsx với dữ liệu từ database

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Chưa cài đặt thư viện openpyxl")
    print("Vui lòng chạy: pip install openpyxl")
    exit(1)

# Lấy danh sách nhân viên từ database
try:
    from services.employee_services import EmployeeService
    employee_service = EmployeeService()
    employees = employee_service.get_all_employees()
    
    if not employees:
        print("⚠️  Cảnh báo: Không có nhân viên nào trong database!")
        print("📝 Hãy thêm nhân viên vào hệ thống trước, sau đó chạy lại script này.")
        print("🔄 Script sẽ tạo file mẫu với dữ liệu giả định...")
        employees = []  # Sẽ dùng dữ liệu mẫu bên dưới
    else:
        print(f"✅ Đã tìm thấy {len(employees)} nhân viên trong database")
        print(f"📊 Sẽ tạo file mẫu với tối đa 10 nhân viên đầu tiên...")
except Exception as e:
    print(f"⚠️  Không thể kết nối database: {e}")
    print("🔄 Sẽ tạo file mẫu với dữ liệu giả định...")
    employees = []

# Tạo workbook mới
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Danh sách nhân viên"

# Định dạng header
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Tạo header
headers = ['Mã nhân viên', 'Tên nhân viên', 'Ghi chú']
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Chuẩn bị dữ liệu
if employees:
    # Dùng dữ liệu thực từ database (lấy tối đa 10 nhân viên)
    data = []
    for emp in employees[:10]:
        data.append([
            emp.get('employee_code', ''),
            emp.get('name', ''),
            f"Phòng: {emp.get('department_id', 'N/A')} - Chức vụ: {emp.get('job_title_id', 'N/A')}"
        ])
    print(f"📋 Đang tạo file với {len(data)} nhân viên từ database...")
else:
    # Dữ liệu mẫu giả định (khi không có nhân viên trong DB)
    data = [
        ['00001', 'Nguyễn Văn A', 'Nhân viên phòng kế toán'],
        ['00002', 'Trần Thị B', 'Nhân viên phòng kinh doanh'],
        ['00003', 'Lê Văn C', 'Nhân viên phòng kỹ thuật'],
        ['00004', 'Phạm Thị D', 'Nhân viên phòng hành chính'],
        ['00005', 'Hoàng Văn E', 'Nhân viên phòng IT'],
        ['00010', 'Võ Thị F', 'Nhân viên phòng nhân sự'],
        ['00015', 'Đặng Văn G', 'Nhân viên phòng marketing'],
        ['00020', 'Bùi Thị H', 'Nhân viên phòng thiết kế'],
        ['00025', 'Dương Văn I', 'Nhân viên phòng bảo vệ'],
        ['00030', 'Đinh Thị K', 'Nhân viên phòng tài chính'],
    ]
    print(f"📋 Đang tạo file với {len(data)} nhân viên mẫu (dữ liệu giả định)...")

# Định dạng dữ liệu
data_font = Font(name='Arial', size=10)
data_alignment = Alignment(horizontal='left', vertical='center')
code_alignment = Alignment(horizontal='center', vertical='center')

# Thêm dữ liệu
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = border
        
        # Căn giữa cho cột mã nhân viên
        if col_idx == 1:
            cell.alignment = code_alignment
        else:
            cell.alignment = data_alignment

# Tự động điều chỉnh độ rộng cột
ws.column_dimensions['A'].width = 15  # Mã nhân viên
ws.column_dimensions['B'].width = 30  # Tên nhân viên
ws.column_dimensions['C'].width = 35  # Ghi chú

# Đóng băng dòng đầu tiên
ws.freeze_panes = 'A2'

# Lưu file
output_file = 'upload_nhanvien_mau.xlsx'
wb.save(output_file)

print(f"\n✅ Đã tạo file {output_file} thành công!")
print(f"📁 File được lưu tại: {output_file}")

if employees:
    print(f"\n✨ File chứa {len(data)} nhân viên THỰC từ database của bạn")
    print(f"📊 Các mã nhân viên có trong file:")
    for emp_data in data[:5]:
        print(f"   - {emp_data[0]}: {emp_data[1]}")
    if len(data) > 5:
        print(f"   ... và {len(data) - 5} nhân viên khác")
else:
    print(f"\n⚠️  File chứa dữ liệu MẪU (không có trong database)")
    print(f"💡 Để tạo file với dữ liệu thực:")
    print(f"   1. Thêm nhân viên vào hệ thống qua menu 'Khai báo > Thông tin nhân viên'")
    print(f"   2. Chạy lại script này: python create_upload_template.py")

print("\n📝 Hướng dẫn sử dụng:")
print("1. Mở file Excel vừa tạo")
print("2. Chỉnh sửa danh sách (thêm/bớt nhân viên) theo nhu cầu")
print("3. Lưu file")
print("4. Trong phần mềm, click nút '📤 Upload danh sách'")
print("5. Chọn file Excel vừa chỉnh sửa")
print("6. Hệ thống sẽ tự động thêm nhân viên vào danh sách tải lên")
